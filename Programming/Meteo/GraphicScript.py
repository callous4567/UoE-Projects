import matplotlib.pyplot as plt
import numpy as np
import time
import random
import sys
import matplotlib.ticker as pltick
from openpyxl import Workbook
from openpyxl import load_workbook

# Load the workbook in question.
wb1 = load_workbook('Obs2018.xlsx')

def iter_rows(ws): # This defines iter_rows(ws), w/ ws being the worksheet in question.
    result = [] # The result, a list of lists, each sublist being a row, [[1,2,3],[4,5,6]...]
    for row in ws.iter_rows(): # Iteration for each row: row is an object that is called for iter_rows()
        rowlist = [] # Each row list [1,2,3]
        for cell in row: # Takes the cells for each row: cell is an object called for the object row.
            rowlist.append(cell.value) # This appends the value of each cell object.
        result.append(rowlist) # Appends the rowlist to the result as a list of lists.
    return result # Returns the list of lists for the document.

def myround(x):
    return round(x, 1)

def standard_deviation(list):
    list_sum = sum(list)
    average = (1/len(list))*list_sum
    dif_square_list = [(d - average)**2 for d in list]
    list_dif_square_sum = sum(dif_square_list)
    sigma = np.sqrt((1/len(list))*list_dif_square_sum)
    return sigma, average




def tuple_sort(list):
    sorted_list = list.sort(key=lambda x: x[1]) # Sorts tuple by second value
    return sorted_list

# We're going to need graphs that have the standard deviation provided.
# Temperature max, min, time_taken, and dewpoint, can all be plotted on one graph.

# Next step is correction for erroneous data. I can't automate this.
# Load the workbook in question.
wb2 = load_workbook('correctedObs2018.xlsx') # Loads corrected value excel, with erroneous data set to "0" (or you could set it to some other letter, I chose 0 for ease)

    
def graph_graph(length, actual, maxi, mini, dew, wet, rel, seap, corrected_sheet):

    # This is the temperature grapher.
    
    fig = plt.figure(111, figsize=(16,12))
    ax = plt.subplot(111)
    ax.set(xlim=[1,len(length)], ylim=[0, 20], xlabel=(r'$\it{Student}$'), ylabel=("Temperature / ($^\circ$C)"), title=("Graph of recorded Temperature variables against Student"))


    ax.plot(length, actual, label="Actual recorded Temperature", color='black', marker='x')
    a_a = []
    for d in range(1, len(length) + 1):
        a_a.append(standard_deviation(corrected_sheet[1])[1])
    ax.plot(length, a_a, label="Mean Actual", color='black', linewidth=1, linestyle="--")
    
    
    ax.plot(length, maxi, label="Maximum Temperature recorded", color='red', marker='x')
    m_m = []
    for d in range(1, len(length) + 1):
        m_m.append(standard_deviation(corrected_sheet[2])[1])
    ax.plot(length, m_m, label="Mean Maximum", color='red', linewidth=1, linestyle="--")
    
    
    ax.plot(length, mini, label="Minimum Temperature recorded", color='blue', marker='x')
    mi_mi = []
    for d in range(1, len(length) + 1):
        mi_mi.append(standard_deviation(corrected_sheet[3])[1])
    ax.plot(length, mi_mi, label="Mean Minimum", color='blue', linewidth=1, linestyle="--")

    
    ax.plot(length, dew, label="Dewpoint Temperature Calculated", color='green', marker='x')
    d_d = []
    for d in range(1, len(length) + 1):
        d_d.append(standard_deviation(corrected_sheet[5])[1])
    ax.plot(length, d_d, label="Mean Dewpoint", color='green', linewidth=1, linestyle="--")

    
    ax.plot(length, wet, label="Wet Bulb Temperature recorded", color='lightblue', marker='x')
    w_w = []
    for d in range(1, len(length) + 1):
        w_w.append(standard_deviation(corrected_sheet[4])[1])
    ax.plot(length, w_w, label="Mean Wet-Bulb", color='lightblue', linewidth=1, linestyle="--")


    ax.text(6, standard_deviation(corrected_sheet[1])[1] + 1, ("{1} Actual equals {0:.1e} and average equals {2:.1f}").format(standard_deviation(corrected_sheet[1])[0], r'$\it\sigma$', myround(standard_deviation(corrected_sheet[1])[1])))
    ax.text(6, standard_deviation(corrected_sheet[2])[1] + 1, ("{1} Maximum equals {0:.1e} and average equals {2:.1f}").format(standard_deviation(corrected_sheet[2])[0], r'$\it\sigma$', myround(standard_deviation(corrected_sheet[2])[1])))
    ax.text(6, standard_deviation(corrected_sheet[3])[1] + 1, ("{1} Minimum equals {0:.1e} and average equals {2:.1f}").format(standard_deviation(corrected_sheet[3])[0], r'$\it\sigma$', myround(standard_deviation(corrected_sheet[3])[1])))
    ax.text(6, standard_deviation(corrected_sheet[5])[1] + 1, ("{1} Wet Bulb equals {0:.1e} and average equals {2:.1f}").format(standard_deviation(corrected_sheet[5])[0], r'$\it\sigma$', myround(standard_deviation(corrected_sheet[5])[1])))
    ax.text(6, standard_deviation(corrected_sheet[4])[1], ("{1} Dew equals {0:.1e} and average equals {2:.1f}").format(standard_deviation(corrected_sheet[4])[0], r'$\it\sigma$', myround(standard_deviation(corrected_sheet[4])[1])))
    
    ax.grid(True, which='major')
    ax.grid(True, which='minor')
    ax.minorticks_on()
    
    majorticks_student = pltick.MultipleLocator(1)
    majorticks_temp = pltick.MultipleLocator(2)
    minorticks_student = pltick.MultipleLocator(1)
    minorticks_temp = pltick.MultipleLocator(0.5)
    
    ax.xaxis.set_major_locator(majorticks_student)
    ax.yaxis.set_major_locator(majorticks_temp)
    ax.xaxis.set_minor_locator(minorticks_student)
    ax.yaxis.set_minor_locator(minorticks_temp)
    
    ax.spines['right'].set_visible(False)
    ax.spines['top'].set_visible(False)
    plt.legend()
    plt.show()
    fig.savefig(("temperature{}.png").format(str(dew[1])))

    # This is the pressure grapher

    fig1 = plt.figure(222, figsize=(16, 6))
    ax1 = plt.subplot(111)

    ax1.set(xlim=[1, len(length)], ylim=[min(seap) - 0.1,max(seap) + 0.1], xlabel="Student", ylabel="Pressure ($\it{hPa}$)", title="Recorded Pressure by Student")
    
    seap_avg = []
    for d in range(0, len(seap)):
        seap_avg.append(standard_deviation(corrected_sheet[7])[1])
        
    ax1.plot(length, seap_avg, color='black', linewidth=0.5, label="Average Pressure recorded")
    ax1.plot(length, seap, label="Pressure as recorded", color='black', marker='x')
    
    ax1.grid(True, which='major')
    ax1.grid(True, which='minor')
    majors = pltick.MultipleLocator(1)
    minors = pltick.MultipleLocator(0.1)
    ax1.xaxis.set_major_locator(majors)
    ax1.xaxis.set_minor_locator(majors)
    ax1.yaxis.set_major_locator(minors)
    ax1.text(1, min(seap), ("{1} Pressure equals {0:.1e}").format(standard_deviation(corrected_sheet[7])[0], "$\it\sigma$"))
    ax1.spines['right'].set_visible(False)
    ax1.spines['top'].set_visible(False)
    ax1.margins(0, 16)
    ax1.text(1, max(seap), ("Average pressure equals {0:.1f}").format(myround(standard_deviation(corrected_sheet[7])[1])))
    
    plt.legend()
    plt.show()
    fig1.savefig(("pressure{}.png").format(str(dew[1])))

    # This is the humidity grapher

    fig2 = plt.figure(333, figsize=(16,6))
    ax2 = plt.subplot(111)

    ax2.set(xlim=[1, len(length)], ylim=[min(rel) - 1, max(rel) + 1], title="Graph of Relative Humidity $\it{RH}$ expressed in % versus Student", xlabel="Student", ylabel="$\it{RH}$ in %")
    ax2.plot(length, rel, label="Relative Humidity recorded", color="black", marker="x")

    avgg = []
    for d in range(0, len(rel)):
        avgg.append(standard_deviation(corrected_sheet[6])[1])
    ax2.plot(length, avgg, label="Average Relative Humidity", color='black', linewidth=0.5)

    ax2.grid(True, which='major', axis='both')
    ax2.grid(True, which='minor', axis='both')
    ax2.text(1, min(rel), ("{1} Relative Humidity equals {0:.0e}").format(standard_deviation(corrected_sheet[6])[0], "$\it\sigma$"))
    ax2.text(1, max(rel) - 4, ("Relative Humidity Average equals {0:d} %").format(round(standard_deviation(corrected_sheet[6])[1]), "$\it\sigma$"))

    ax2.minorticks_on()
    ax2.xaxis.set_major_locator(pltick.MultipleLocator(1))
    ax2.xaxis.set_minor_locator(pltick.MultipleLocator(1))
    ax2.yaxis.set_minor_locator(pltick.MultipleLocator(1))
    ax2.yaxis.set_major_locator(pltick.MultipleLocator(1))
    ax2.spines['right'].set_visible(False)
    ax2.spines['top'].set_visible(False)
    ax2.margins(0,0)
    ax2.spines['bottom'].set_visible(True)
    plt.legend()
    
    plt.show()
    fig2.savefig(("relativehumidity{}.png").format(str(dew[1])))



            
# Now to do the pressure and relative humidity graphs.
# ws.iter_rows() gets you raw data which can be called by the 'row' and then by the 'cell' of each row. I.e. for a row 4 and cell E, you'd get 4:E. You can't directly use ws.iter_rows() for formatting, so you need to convert it into a list of lists, each list for each row. That's what this does.
# So, we have a list of lists... we need to isolate the elements of each column.
# Starting row: 4
# Columns: 1: Student N0, 2: Air Temp, 3: Max Temp, 4: Min Temp, 5: Dew Point, 6: Relative Hum, 7: Mean Sea Level Pressure
# Let's do this function-wise.

wb2 = load_workbook("correctedObs2018.xlsx")

def erroneous_corrector(sheet_label):
    sheet = wb2[sheet_label]
    list_ol = iter_rows(sheet) # Gets you a list of lists, of all elements
    new_list_ol = list_ol[4:len(list_ol)] # Cuts off first four rows, i.e. starts from 4. list_ol[0,1,2,3] are all bad elements, 4 is row 5.
    length = []
    for d in range(0, len(new_list_ol)):
        length.append(d + 1) # Marks the number in this length list

    dry = []
    maxi = []
    mini = []
    wet = []
    dew = []
    rel = []
    seap = []

    for col in new_list_ol: # Calls the columns of the list.

        dry.append(col[1]) # First column is the dry.
        maxi.append(col[2])
        mini.append(col[3])
        wet.append(col[4])
        dew.append(col[5])
        rel.append(col[6])
        seap.append(col[7])

    return length, dry, maxi, mini, dew, wet, rel, seap

# We have all the lists, uncorrected. Now, to correct.
def zero_remover(listo):
    new = []
    for d in listo: # It's a nested list. So treat each d as a list of its own.
        d = list(filter((0).__ne__, d)) # Use list notation.
        new.append(d)
    return(new)


def sheet_tabulator(sheet_label):
    sheet = wb1[sheet_label]
    max_row = sheet.max_row
    max_column = sheet.max_column
    list_ol = iter_rows(sheet) # We need to isolate off the first four rows.
    new_list_ol = list_ol[4:len(list_ol)] # Selects from row 5 (4+1) to length.
    
    name = [] # Fucked up (A5 + 1) bullshit.
    
    realname = []
    for d in range(0,len(new_list_ol)):
        realname.append(d + 1)
    
    dry = []
    maxi = []
    mini = []
    wet = []
    dew = []
    rel = []
    seap = []

    for col in new_list_ol: # Calls the columns of the list of lists... i.e. when stacked,
                            # The element of each column [0] is the start of it... 
        name.append(col[0])
        dry.append(col[1])
        maxi.append(col[2])
        mini.append(col[3])
        wet.append(col[4])
        dew.append(col[5])
        rel.append(col[6])
        seap.append(col[7])


    new_sheet = erroneous_corrector(sheet_label)
    zeroed_sheet = zero_remover(new_sheet)
    corrected_sheet = zeroed_sheet
    print(zeroed_sheet[7])
    graph_graph(realname, dry, maxi, mini, dew, wet, rel, seap, corrected_sheet) # Incase you need to graph it using our grapher.
    

    return realname, dry, maxi, mini, dew, wet, rel, seap

sheet_tabulator("Monday")
sheet_tabulator("Tuesday")
sheet_tabulator("Thursday")
